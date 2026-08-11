import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.layout import Layout, ManualLayout


# =========================================================
# COLUMN NAME DETECTION
# =========================================================

COLUMN_ALIASES = {
    "product": ["product", "product name", "item", "item name"],
    "quantity": ["quantity", "qty", "units", "units sold"],
    "price": ["price", "unit price", "selling price", "unitprice"],
    "salesperson": ["salesperson", "sales person", "rep", "sales rep", "employee"],
    "region": ["region", "area", "zone", "territory"],
    "sales": ["sales", "total sales", "amount", "revenue", "total"],
}


def normalize(header):
    return " ".join(header.strip().lower().split())


def build_header_map(fieldnames):
    normalized_to_actual = {
        normalize(h): h
        for h in fieldnames
    }

    header_map = {}

    for concept, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized_to_actual:
                header_map[concept] = normalized_to_actual[alias]
                break

    return header_map


def to_number(value, row_number, column_name):
    try:
        return float(
            str(value).replace(",", "").strip()
        )
    except (TypeError, ValueError):
        raise ValueError(
            f"Could not convert value '{value}' in column "
            f"'{column_name}' (data row {row_number}) to a number."
        )


# =========================================================
# GENERATE REPORT
# =========================================================

def generate_report():

    # -----------------------------------------------------
    # FILE PATHS
    # -----------------------------------------------------

    base_dir = Path(__file__).resolve().parent

    input_file = (
        base_dir
        / "data"
        / "sales_data.csv"
    )

    output_file = (
        base_dir
        / "sales_report.xlsx"
    )

    # -----------------------------------------------------
    # CHECK INPUT FILE
    # -----------------------------------------------------

    if not input_file.exists():
        raise FileNotFoundError(
            f"sales_data.csv was not found here:\n{input_file}"
        )

    # -----------------------------------------------------
    # READ CSV
    # -----------------------------------------------------

    with input_file.open(
        mode="r",
        encoding="utf-8-sig",
        newline=""
    ) as file:

        reader = csv.DictReader(file)

        if not reader.fieldnames:
            raise ValueError(
                f"The CSV file at {input_file} appears to have no headers."
            )

        raw_rows = list(reader)

    if not raw_rows:
        raise ValueError(
            f"The CSV file at {input_file} has headers but no data rows."
        )

    # -----------------------------------------------------
    # DETECT COLUMNS
    # -----------------------------------------------------

    header_map = build_header_map(
        reader.fieldnames
    )

    required_always = [
        "product",
        "salesperson",
        "region"
    ]

    missing_always = [
        column
        for column in required_always
        if column not in header_map
    ]

    has_sales = "sales" in header_map

    has_qty_and_price = (
        "quantity" in header_map
        and "price" in header_map
    )

    if missing_always or not (
        has_sales or has_qty_and_price
    ):

        found = ", ".join(
            reader.fieldnames
        )

        problems = []

        if missing_always:
            problems.append(
                "missing required column(s) for: "
                + ", ".join(missing_always)
            )

        if not (
            has_sales or has_qty_and_price
        ):
            problems.append(
                "no 'Sales' column found, and not enough "
                "of 'Quantity' + 'Price/Unit Price' to calculate it"
            )

        raise ValueError(
            "The CSV is missing required business data.\n"
            f"Problem(s): {'; '.join(problems)}.\n"
            f"Columns actually found in the CSV: {found}"
        )

    # -----------------------------------------------------
    # DETERMINE HOW SALES WILL BE CALCULATED
    # -----------------------------------------------------

    calculate_sales = not has_sales

    # -----------------------------------------------------
    # CLEAN AND PREPARE DATA
    # -----------------------------------------------------

    sales_data = []

    for row_number, row in enumerate(
        raw_rows,
        start=2
    ):

        product = row[
            header_map["product"]
        ].strip()

        salesperson = row[
            header_map["salesperson"]
        ].strip()

        region = row[
            header_map["region"]
        ].strip()

        quantity = None
        price = None

        if "quantity" in header_map:
            quantity = to_number(
                row[header_map["quantity"]],
                row_number,
                header_map["quantity"]
            )

        if "price" in header_map:
            price = to_number(
                row[header_map["price"]],
                row_number,
                header_map["price"]
            )

        if calculate_sales:

            sales = quantity * price

        else:

            sales = to_number(
                row[header_map["sales"]],
                row_number,
                header_map["sales"]
            )

        sales_data.append(
            {
                "Product": product,
                "Quantity": quantity,
                "Price": price,
                "Salesperson": salesperson,
                "Region": region,
                "Sales": sales,
            }
        )

    # =====================================================
    # CALCULATE KEY METRICS
    # =====================================================

    total_sales = sum(
        row["Sales"]
        for row in sales_data
    )

    total_quantity = sum(
        row["Quantity"]
        for row in sales_data
        if row["Quantity"] is not None
    )

    transactions = len(
        sales_data
    )

    average_sale = (
        total_sales / transactions
        if transactions
        else 0
    )

    # -----------------------------------------------------
    # PRODUCT SALES
    # -----------------------------------------------------

    product_sales = {}

    for row in sales_data:

        product = row["Product"]

        product_sales[product] = (
            product_sales.get(product, 0)
            + row["Sales"]
        )

    top_product = max(
        product_sales,
        key=product_sales.get
    )

    top_product_sales = product_sales[
        top_product
    ]

    # -----------------------------------------------------
    # SALESPERSON SALES
    # -----------------------------------------------------

    salesperson_sales = {}

    for row in sales_data:

        salesperson = row["Salesperson"]

        salesperson_sales[salesperson] = (
            salesperson_sales.get(
                salesperson,
                0
            )
            + row["Sales"]
        )

    best_salesperson = max(
        salesperson_sales,
        key=salesperson_sales.get
    )

    best_salesperson_sales = (
        salesperson_sales[
            best_salesperson
        ]
    )

    # -----------------------------------------------------
    # REGION SALES
    # -----------------------------------------------------

    region_sales = {}

    for row in sales_data:

        region = row["Region"]

        region_sales[region] = (
            region_sales.get(
                region,
                0
            )
            + row["Sales"]
        )

    best_region = max(
        region_sales,
        key=region_sales.get
    )

    best_region_sales = (
        region_sales[
            best_region
        ]
    )

    # =====================================================
    # CREATE WORKBOOK
    # =====================================================

    workbook = Workbook()

    sales_report = workbook.active
    sales_report.title = "Sales Report"

    summary = workbook.create_sheet(
        "Summary"
    )

    # =====================================================
    # STYLES
    # =====================================================

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="5B9BD5"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    thin_side = Side(
        style="thin",
        color="B7B7B7"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    # =====================================================
    # SALES REPORT SHEET
    # =====================================================

    sales_report.merge_cells(
        "A1:F1"
    )

    sales_report["A1"] = (
        "SALES REPORT"
    )

    sales_report["A1"].font = (
        title_font
    )

    sales_report["A1"].fill = (
        title_fill
    )

    sales_report["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    headers = [
        "Product",
        "Quantity",
        "Price",
        "Salesperson",
        "Region",
        "Sales"
    ]

    for column, header in enumerate(
        headers,
        start=1
    ):

        cell = sales_report.cell(
            row=3,
            column=column,
            value=header
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for row_number, row in enumerate(
        sales_data,
        start=4
    ):

        values = [
            row["Product"],
            row["Quantity"],
            row["Price"],
            row["Salesperson"],
            row["Region"],
            row["Sales"]
        ]

        for column, value in enumerate(
            values,
            start=1
        ):

            cell = sales_report.cell(
                row=row_number,
                column=column,
                value=value
            )

            cell.border = border

            cell.alignment = Alignment(
                vertical="center"
            )

        sales_report.cell(
            row=row_number,
            column=3
        ).number_format = (
            '₹#,##0.00'
        )

        sales_report.cell(
            row=row_number,
            column=6
        ).number_format = (
            '₹#,##0.00'
        )

    sales_report.column_dimensions[
        "A"
    ].width = 18

    sales_report.column_dimensions[
        "B"
    ].width = 12

    sales_report.column_dimensions[
        "C"
    ].width = 15

    sales_report.column_dimensions[
        "D"
    ].width = 18

    sales_report.column_dimensions[
        "E"
    ].width = 15

    sales_report.column_dimensions[
        "F"
    ].width = 15

    sales_report.freeze_panes = "A4"

    # =====================================================
    # SUMMARY SHEET
    # =====================================================

    summary.merge_cells(
        "A1:H1"
    )

    summary["A1"] = (
        "SALES PERFORMANCE SUMMARY"
    )

    summary["A1"].font = (
        title_font
    )

    summary["A1"].fill = (
        title_fill
    )

    summary["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    summary.row_dimensions[
        1
    ].height = 30

    # -----------------------------------------------------
    # SECTION TITLES
    # -----------------------------------------------------

    summary["A3"] = (
        "Product Performance"
    )

    summary["D3"] = (
        "Salesperson Performance"
    )

    summary["G3"] = (
        "Regional Performance"
    )

    for cell in [
        summary["A3"],
        summary["D3"],
        summary["G3"]
    ]:

        cell.font = Font(
            bold=True,
            size=14,
            color="5B5B5B"
        )

    # -----------------------------------------------------
    # TABLE HEADERS
    # -----------------------------------------------------

    table_headers = {
        "A4": "Product",
        "B4": "Total Sales",
        "D4": "Salesperson",
        "E4": "Total Sales",
        "G4": "Region",
        "H4": "Total Sales"
    }

    for cell_address, value in (
        table_headers.items()
    ):

        cell = summary[cell_address]

        cell.value = value
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =====================================================
    # PRODUCT TABLE
    # =====================================================

    product_start_row = 5

    row_number = product_start_row

    for product, sales in (
        product_sales.items()
    ):

        cell = summary.cell(
            row=row_number,
            column=1,
            value=product
        )

        cell.border = border

        cell.alignment = Alignment(
            vertical="center"
        )

        cell = summary.cell(
            row=row_number,
            column=2,
            value=sales
        )

        cell.border = border

        cell.number_format = (
            '₹#,##0.00'
        )

        cell.alignment = Alignment(
            vertical="center"
        )

        row_number += 1

    product_end_row = (
        row_number - 1
    )

    # =====================================================
    # SALESPERSON TABLE
    # =====================================================

    salesperson_start_row = 5

    row_number = salesperson_start_row

    for salesperson, sales in (
        salesperson_sales.items()
    ):

        cell = summary.cell(
            row=row_number,
            column=4,
            value=salesperson
        )

        cell.border = border

        cell.alignment = Alignment(
            vertical="center"
        )

        cell = summary.cell(
            row=row_number,
            column=5,
            value=sales
        )

        cell.border = border

        cell.number_format = (
            '₹#,##0.00'
        )

        cell.alignment = Alignment(
            vertical="center"
        )

        row_number += 1

    salesperson_end_row = (
        row_number - 1
    )

    # =====================================================
    # REGION TABLE
    # =====================================================

    region_start_row = 5

    row_number = region_start_row

    for region, sales in (
        region_sales.items()
    ):

        cell = summary.cell(
            row=row_number,
            column=7,
            value=region
        )

        cell.border = border

        cell.alignment = Alignment(
            vertical="center"
        )

        cell = summary.cell(
            row=row_number,
            column=8,
            value=sales
        )

        cell.border = border

        cell.number_format = (
            '₹#,##0.00'
        )

        cell.alignment = Alignment(
            vertical="center"
        )

        row_number += 1

    region_end_row = (
        row_number - 1
    )

    max_table_end_row = max(
        product_end_row,
        salesperson_end_row,
        region_end_row
    )

    # =====================================================
    # KEY METRICS
    # =====================================================

    kpi_title_row = (
        max_table_end_row + 2
    )

    summary.cell(
        row=kpi_title_row,
        column=1,
        value="Key Metrics"
    ).font = Font(
        bold=True,
        size=14,
        color="5B5B5B"
    )

    kpi_header_row = (
        kpi_title_row + 1
    )

    summary.cell(
        row=kpi_header_row,
        column=1,
        value="Metric"
    )

    summary.cell(
        row=kpi_header_row,
        column=2,
        value="Value"
    )

    for column in [1, 2]:

        cell = summary.cell(
            row=kpi_header_row,
            column=column
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    kpi_data = [
        ("Total Sales", total_sales),
        ("Total Quantity", total_quantity),
        ("Transactions", transactions),
        ("Average Sale", average_sale),
        ("Top Product", top_product),
        ("Top Product Sales", top_product_sales),
        ("Best Salesperson", best_salesperson),
        ("Best Salesperson Sales", best_salesperson_sales),
        ("Best Region", best_region),
        ("Best Region Sales", best_region_sales)
    ]

    kpi_start_row = (
        kpi_header_row + 1
    )

    currency_indices = {
        0,
        3,
        5,
        7,
        9
    }

    for idx, (metric, value) in enumerate(
        kpi_data
    ):

        row_num = (
            kpi_start_row + idx
        )

        metric_cell = summary.cell(
            row=row_num,
            column=1,
            value=metric
        )

        metric_cell.border = border

        metric_cell.alignment = Alignment(
            vertical="center"
        )

        value_cell = summary.cell(
            row=row_num,
            column=2,
            value=value
        )

        value_cell.border = border

        value_cell.alignment = Alignment(
            vertical="center"
        )

        if idx in currency_indices:

            value_cell.number_format = (
                '₹#,##0.00'
            )

    kpi_end_row = (
        kpi_start_row
        + len(kpi_data)
        - 1
    )

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    summary.column_dimensions[
        "A"
    ].width = 24

    summary.column_dimensions[
        "B"
    ].width = 18

    summary.column_dimensions[
        "C"
    ].width = 4

    summary.column_dimensions[
        "D"
    ].width = 24

    summary.column_dimensions[
        "E"
    ].width = 18

    summary.column_dimensions[
        "F"
    ].width = 4

    summary.column_dimensions[
        "G"
    ].width = 20

    summary.column_dimensions[
        "H"
    ].width = 18

    # =====================================================
    # CHART FUNCTION
    # =====================================================

    def add_bar_chart(
        anchor_cell,
        title,
        x_title,
        cat_col,
        data_col,
        data_start_row,
        data_end_row
    ):

        chart = BarChart()

        chart.type = "col"
        chart.style = 10
        chart.title = title

        data = Reference(
            summary,
            min_col=data_col,
            min_row=data_start_row,
            max_row=data_end_row
        )

        categories = Reference(
            summary,
            min_col=cat_col,
            min_row=data_start_row,
            max_row=data_end_row
        )

        chart.add_data(
            data,
            titles_from_data=False
        )

        chart.set_categories(
            categories
        )

        chart.series[0].tx = (
            SeriesLabel(v="Sales")
        )

        chart.legend = None

        chart.x_axis.title = (
            x_title
        )

        chart.y_axis.title = (
            "Sales"
        )

        chart.height = 14
        chart.width = 17

        chart.plot_area.layout = Layout(
            manualLayout=ManualLayout(
                x=0.12,
                y=0.10,
                w=0.82,
                h=0.52,
                xMode="edge",
                yMode="edge"
            )
        )

        chart.x_axis.delete = False
        chart.y_axis.delete = False

        chart.x_axis.tickLblPos = "low"

        summary.add_chart(
            chart,
            anchor_cell
        )

    # =====================================================
    # ADD CHARTS
    # =====================================================

    chart_start_row = (
        kpi_end_row + 3
    )

    # Chart 1 - Product
    add_bar_chart(
        f"A{chart_start_row}",
        "Sales by Product",
        "Product",
        cat_col=1,
        data_col=2,
        data_start_row=product_start_row,
        data_end_row=product_end_row
    )

    # Chart 2 - Salesperson
    add_bar_chart(
        f"A{chart_start_row + 28}",
        "Sales by Salesperson",
        "Salesperson",
        cat_col=4,
        data_col=5,
        data_start_row=salesperson_start_row,
        data_end_row=salesperson_end_row
    )

    # Chart 3 - Region
    add_bar_chart(
        f"A{chart_start_row + 56}",
        "Sales by Region",
        "Region",
        cat_col=7,
        data_col=8,
        data_start_row=region_start_row,
        data_end_row=region_end_row
    )

    # =====================================================
    # FREEZE PANES
    # =====================================================

    summary.freeze_panes = "A4"

    # =====================================================
    # SAVE WORKBOOK
    # =====================================================

    workbook.save(
        output_file
    )

    # =====================================================
    # SUCCESS MESSAGE
    # =====================================================

    print()
    print(
        "======================================"
    )
    print(
        "       SALES REPORT GENERATED"
    )
    print(
        "======================================"
    )

    print(
        f"Input file : {input_file}"
    )

    print(
        f"Output file: {output_file}"
    )

    print()

    print(
        "Report generated successfully!"
    )


# =========================================================
# DIRECT EXECUTION
# =========================================================

if __name__ == "__main__":
    generate_report()